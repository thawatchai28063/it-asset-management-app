from __future__ import annotations

from datetime import datetime
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\it-ae\Desktop\IT ASSET TRACKING Control.xlsx")
OUTPUT = ROOT / "tools" / "import_asset_tracking_control.sql"


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sql(value: str | None) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def date_value(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    try:
        parsed = datetime.fromisoformat(str(value).replace(" 00:00:00", ""))
        return parsed.strftime("%Y-%m-%d")
    except ValueError:
        return None


def status_value(value: str) -> str:
    normalized = value.strip().lower()
    if normalized == "used":
        return "in_use"
    if normalized in {"repair", "maintenance"}:
        return "repair"
    if normalized in {"retired", "scrap", "broken"}:
        return "retired"
    return "available"


def note_value(remark: str, password_admin: str, check_date_raw: str) -> str | None:
    parts = []
    if remark:
        parts.append(f"remark: {remark}")
    if password_admin:
        parts.append(f"passwordadmin: {password_admin}")
    if check_date_raw and date_value(check_date_raw) is None:
        parts.append(f"check_date_raw: {check_date_raw}")
    return "\n".join(parts) if parts else None


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    headers = [text(cell.value) for cell in sheet[2]]
    col = {name: index for index, name in enumerate(headers)}
    excel_rows = list(sheet.iter_rows(min_row=3, values_only=True))
    serial_counts = Counter(
        text(excel_row[col["Serialnumber"]])
        for excel_row in excel_rows
        if text(excel_row[col["Serialnumber"]])
    )

    rows: list[dict[str, str | None]] = []
    for excel_row in excel_rows:
        it_tag = text(excel_row[col["IT Tag"]])
        if not it_tag:
            continue

        brand = text(excel_row[col["Band"]])
        model = text(excel_row[col["Model"]])
        asset_type = text(excel_row[col["TYPE"]]) or "Other"
        raw_serial = text(excel_row[col["Serialnumber"]])
        serial = raw_serial
        if not serial or serial == "-" or serial_counts[raw_serial] > 1:
            serial = it_tag
        department = text(excel_row[col["DEPARTMENT/USAGE"]]) or "Unassigned"
        assigned_user = text(excel_row[col["NAME"]]) or None
        position = text(excel_row[col["Location"]]) or None
        purchase_date = date_value(excel_row[col["Installed"]])
        note = note_value(
            text(excel_row[col["remark"]]),
            text(excel_row[col["Password admin"]]),
            text(excel_row[col["Check Date"]]),
        )

        asset_name = " ".join(part for part in [brand, model] if part).strip()
        if not asset_name:
            asset_name = f"{asset_type} {it_tag}"

        rows.append(
            {
                "asset_name": asset_name,
                "asset_type": asset_type,
                "it_tag": it_tag,
                "employee_no": text(excel_row[col["Emp,No."]]) or None,
                "description": text(excel_row[col["DESCRIPTION"]]) or None,
                "os_version": text(excel_row[col["OS/VERSION"]]) or None,
                "brand": brand or None,
                "model": model or None,
                "serial_number": serial,
                "ip_address": None,
                "department": department,
                "status": status_value(text(excel_row[col["STATUS"]])),
                "assigned_user": assigned_user,
                "position": position,
                "point_image": text(excel_row[col["Point/Image"]]) or None,
                "check_date": date_value(excel_row[col["Check Date"]]),
                "purchase_date": purchase_date,
                "note": note,
            }
        )

    columns = [
        "asset_name",
        "asset_type",
        "it_tag",
        "employee_no",
        "description",
        "os_version",
        "brand",
        "model",
        "serial_number",
        "ip_address",
        "department",
        "status",
        "assigned_user",
        "position",
        "point_image",
        "check_date",
        "purchase_date",
        "note",
    ]

    statements = [
        "USE it_asset_management;",
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
    ]

    for row in rows:
        values = ", ".join(sql(row[column]) for column in columns)
        updates = ", ".join(
            f"{column}=VALUES({column})"
            for column in columns
            if column not in {"it_tag"}
        )
        statements.append(
            "INSERT INTO assets "
            f"({', '.join(columns)}) VALUES ({values}) "
            f"ON DUPLICATE KEY UPDATE {updates};"
        )

    statements.extend(["COMMIT;", f"SELECT COUNT(*) AS imported_rows FROM assets WHERE it_tag LIKE 'IT25%';"])
    OUTPUT.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"rows={len(rows)}")
    print(OUTPUT)


if __name__ == "__main__":
    main()
