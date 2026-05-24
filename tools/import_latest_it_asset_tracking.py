from __future__ import annotations

from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "IT Asset Control_repaired.xlsx"
OUTPUT = ROOT / "tools" / "import_latest_it_asset_tracking.sql"


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def date_value(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
    raw = str(value).strip()
    try:
        return datetime.fromisoformat(raw.replace(" 00:00:00", "")).strftime("%Y-%m-%d")
    except ValueError:
        return None


def sql(value: str | None) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def status_value(value: str) -> str:
    normalized = value.strip().lower()
    if normalized == "used":
        return "in_use"
    if normalized in {"repair", "maintenance"}:
        return "repair"
    if normalized in {"retired", "scrap", "broken"}:
        return "retired"
    return "available"


def note_value(row: dict[str, object]) -> str | None:
    parts = []
    for key in ["remark", "Password admin"]:
        value = text(row.get(key))
        if value:
            label = "passwordadmin" if key == "Password admin" else key
            parts.append(f"{label}: {value}")
    return "\n".join(parts) if parts else None


def row_dict(headers: list[str], values: tuple[object, ...]) -> dict[str, object]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["IT Asset Tracking"]
    headers = [text(cell.value) for cell in sheet[2]]
    excel_rows = [row_dict(headers, values) for values in sheet.iter_rows(min_row=3, values_only=True)]

    serial_counts = Counter(
        text(row.get("Serialnumber"))
        for row in excel_rows
        if text(row.get("Serialnumber"))
    )

    columns = [
        "asset_name",
        "asset_type",
        "it_tag",
        "employee_no",
        "description",
        "os_version",
        "brand",
        "model",
        "serial_number",
        "ip_address",
        "department",
        "status",
        "assigned_user",
        "position",
        "point_image",
        "check_date",
        "receipt_of_device",
        "invoice_no",
        "date2",
        "vendor",
        "checker_2025_03_31",
        "check_result_2025_03_31",
        "checker_2025_04_23",
        "checker_2025_05_30",
        "check_result_2025_04_30",
        "purchase_date",
        "note",
    ]

    rows = []
    for row in excel_rows:
        it_tag = text(row.get("IT Tag"))
        if not it_tag:
            continue

        brand = text(row.get("Band"))
        model = text(row.get("Model"))
        asset_type = text(row.get("TYPE")) or "Other"
        raw_serial = text(row.get("Serialnumber"))
        serial = raw_serial
        if not serial or serial == "-" or serial_counts[raw_serial] > 1:
            serial = it_tag

        asset_name = " ".join(part for part in [brand, model] if part).strip()
        if not asset_name:
            asset_name = f"{asset_type} {it_tag}"

        rows.append(
            {
                "asset_name": asset_name,
                "asset_type": asset_type,
                "it_tag": it_tag,
                "employee_no": text(row.get("Emp,No.")) or None,
                "description": text(row.get("DESCRIPTION")) or None,
                "os_version": text(row.get("OS/VERSION")) or None,
                "brand": brand or None,
                "model": model or None,
                "serial_number": serial,
                "ip_address": None,
                "department": text(row.get("DEPARTMENT/USAGE")) or "Unassigned",
                "status": status_value(text(row.get("STATUS"))),
                "assigned_user": text(row.get("NAME")) or None,
                "position": text(row.get("Location")) or None,
                "point_image": text(row.get("Point/Image")) or None,
                "check_date": None,
                "receipt_of_device": date_value(row.get("Receipt of device")),
                "invoice_no": text(row.get("Invoice No.")) or None,
                "date2": date_value(row.get("DATE2")),
                "vendor": text(row.get("VENDOR")) or None,
                "checker_2025_03_31": text(row.get("25/3/31 checker")) or None,
                "check_result_2025_03_31": text(row.get("25/3/31 check result")) or None,
                "checker_2025_04_23": text(row.get("23/4/2025 checker")) or None,
                "checker_2025_05_30": text(row.get("30/5/2025 checker")) or None,
                "check_result_2025_04_30": text(row.get("30/4/2025 check result")) or None,
                "purchase_date": date_value(row.get("Installed")),
                "note": note_value(row),
            }
        )

    statements = [
        "USE it_asset_management;",
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
    ]

    for row in rows:
        values = ", ".join(sql(row[column]) for column in columns)
        updates = ", ".join(
            f"{column}=VALUES({column})"
            for column in columns
            if column != "it_tag"
        )
        statements.append(
            "INSERT INTO assets "
            f"({', '.join(columns)}) VALUES ({values}) "
            f"ON DUPLICATE KEY UPDATE {updates};"
        )

    statements.extend(
        [
            "COMMIT;",
            "SELECT COUNT(*) AS imported_rows FROM assets WHERE it_tag LIKE 'IT25%';",
        ]
    )
    OUTPUT.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"rows={len(rows)}")
    print(OUTPUT)


if __name__ == "__main__":
    main()
