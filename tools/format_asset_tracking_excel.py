from __future__ import annotations

import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SOURCE = Path(r"C:\Users\it-ae\Desktop\IT ASSET TRACKING Control.xlsx")
BACKUP = SOURCE.with_name(
    f"{SOURCE.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{SOURCE.suffix}"
)
CLEANED = SOURCE.with_name(f"{SOURCE.stem}_cleaned{SOURCE.suffix}")
CLEAN_SHEET = "Assets_Import_Export"
README_SHEET = "README"
RAW_SHEET = "Sheet1"


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def date_text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    raw = str(value).strip()
    try:
        return datetime.fromisoformat(raw.replace(" 00:00:00", "")).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def status_value(value: str) -> str:
    normalized = value.strip().lower()
    if normalized == "used":
        return "in_use"
    if normalized in {"repair", "maintenance"}:
        return "repair"
    if normalized in {"retired", "scrap", "broken"}:
        return "retired"
    return "available"


def note_value(remark: str, password_admin: str, check_date_raw: str) -> str:
    parts = []
    if remark:
        parts.append(f"remark: {remark}")
    if password_admin:
        parts.append(f"passwordadmin: {password_admin}")
    if check_date_raw and not date_text(check_date_raw):
        parts.append(f"check_date_raw: {check_date_raw}")
    return "\n".join(parts)


def set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_table_area(ws, max_row: int, max_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    sub_fill = PatternFill("solid", fgColor="E0F2FE")
    required_fill = PatternFill("solid", fgColor="DBEAFE")
    white = "FFFFFF"
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    required_headers = {
        "asset_name",
        "asset_type",
        "it_tag",
        "serial_number",
        "department",
        "status",
    }
    for cell in ws[1]:
        if cell.value in required_headers:
            cell.fill = required_fill
            cell.font = Font(color="0F172A", bold=True)
        elif cell.value:
            cell.fill = sub_fill
            cell.font = Font(color="0F172A", bold=True)


def main() -> None:
    shutil.copy2(SOURCE, BACKUP)
    workbook = load_workbook(SOURCE)
    raw = workbook[RAW_SHEET]
    raw_headers = [text(cell.value) for cell in raw[2]]
    col = {name: index for index, name in enumerate(raw_headers)}
    raw_rows = list(raw.iter_rows(min_row=3, values_only=True))

    serial_counts = Counter(
        text(row[col["Serialnumber"]])
        for row in raw_rows
        if text(row[col["Serialnumber"]])
    )

    if CLEAN_SHEET in workbook.sheetnames:
        del workbook[CLEAN_SHEET]
    if README_SHEET in workbook.sheetnames:
        del workbook[README_SHEET]
    ws = workbook.create_sheet(CLEAN_SHEET, 0)
    readme = workbook.create_sheet(README_SHEET, 1)

    headers = [
        "asset_name",
        "asset_type",
        "it_tag",
        "employee_no",
        "assigned_user",
        "description",
        "brand",
        "model",
        "os_version",
        "serial_number",
        "ip_address",
        "department",
        "position",
        "status",
        "point_image",
        "purchase_date",
        "check_date",
        "note",
        "remark",
        "passwordadmin",
        "original_status",
        "source_item_no",
    ]
    ws.append(headers)

    for raw_row in raw_rows:
        it_tag = text(raw_row[col["IT Tag"]])
        if not it_tag:
            continue

        brand = text(raw_row[col["Band"]])
        model = text(raw_row[col["Model"]])
        asset_type = text(raw_row[col["TYPE"]]) or "Other"
        raw_serial = text(raw_row[col["Serialnumber"]])
        serial_number = raw_serial
        if not serial_number or serial_number == "-" or serial_counts[raw_serial] > 1:
            serial_number = it_tag
        asset_name = " ".join(part for part in [brand, model] if part).strip()
        if not asset_name:
            asset_name = f"{asset_type} {it_tag}"

        remark = text(raw_row[col["remark"]])
        password = text(raw_row[col["Password admin"]])
        check_raw = text(raw_row[col["Check Date"]])

        ws.append(
            [
                asset_name,
                asset_type,
                it_tag,
                text(raw_row[col["Emp,No."]]),
                text(raw_row[col["NAME"]]),
                text(raw_row[col["DESCRIPTION"]]),
                brand,
                model,
                text(raw_row[col["OS/VERSION"]]),
                serial_number,
                "",
                text(raw_row[col["DEPARTMENT/USAGE"]]) or "Unassigned",
                text(raw_row[col["Location"]]),
                status_value(text(raw_row[col["STATUS"]])),
                text(raw_row[col["Point/Image"]]),
                date_text(raw_row[col["Installed"]]),
                date_text(raw_row[col["Check Date"]]),
                note_value(remark, password, check_raw),
                remark,
                password,
                text(raw_row[col["STATUS"]]),
                text(raw_row[col["ITEM NO."]]),
            ]
        )

    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    table = Table(displayName="AssetsImportExport", ref=f"A1:{get_column_letter(max_col)}{max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    style_table_area(ws, max_row, max_col)
    set_widths(
        ws,
        {
            "A": 34,
            "B": 15,
            "C": 13,
            "D": 12,
            "E": 26,
            "F": 22,
            "G": 15,
            "H": 34,
            "I": 20,
            "J": 18,
            "K": 16,
            "L": 22,
            "M": 20,
            "N": 14,
            "O": 14,
            "P": 14,
            "Q": 14,
            "R": 42,
            "S": 18,
            "T": 22,
            "U": 15,
            "V": 12,
        },
    )

    status_col = headers.index("status") + 1
    status_letter = get_column_letter(status_col)
    status_validation = DataValidation(
        type="list",
        formula1='"available,in_use,repair,retired"',
        allow_blank=False,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f"{status_letter}2:{status_letter}{max_row}")

    red_fill = PatternFill("solid", fgColor="FEE2E2")
    duplicate_rule = FormulaRule(
        formula=[f'COUNTIF($C:$C,$C2)>1'],
        fill=red_fill,
    )
    ws.conditional_formatting.add(f"C2:C{max_row}", duplicate_rule)

    readme.sheet_view.showGridLines = False
    readme["A1"] = "IT Asset Import / Export Guide"
    readme["A1"].font = Font(size=16, bold=True, color="1D4ED8")
    readme["A3"] = "Use sheet"
    readme["B3"] = CLEAN_SHEET
    readme["A4"] = "Required columns"
    readme["B4"] = "asset_name, asset_type, it_tag, serial_number, department, status"
    readme["A5"] = "Status values"
    readme["B5"] = "available, in_use, repair, retired"
    readme["A6"] = "Note rule"
    readme["B6"] = "remark and passwordadmin are preserved separately and combined into note for the app."
    readme["A7"] = "Serial rule"
    readme["B7"] = "If serial is blank, '-' or duplicated, use IT Tag as serial_number."
    readme["A8"] = "Original data"
    readme["B8"] = "Sheet1 is kept as the raw source from the original tracking file."
    for row in readme.iter_rows(min_row=3, max_row=8, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="E2E8F0"))
        row[0].font = Font(bold=True)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 90

    # Lightly tidy the original sheet so it remains readable without changing data.
    raw.freeze_panes = "A3"
    raw.auto_filter.ref = f"A2:{get_column_letter(raw.max_column)}{raw.max_row}"
    raw.sheet_view.showGridLines = False
    for cell in raw[1]:
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
    for cell in raw[2]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    raw.row_dimensions[2].height = 30
    for idx in range(1, raw.max_column + 1):
        raw.column_dimensions[get_column_letter(idx)].width = min(28, max(10, len(text(raw.cell(2, idx).value)) + 3))

    try:
        workbook.save(SOURCE)
        saved_path = SOURCE
    except PermissionError:
        workbook.save(CLEANED)
        saved_path = CLEANED
    print(f"saved={saved_path}")
    print(f"backup={BACKUP}")
    print(f"rows={max_row - 1}")


if __name__ == "__main__":
    main()
