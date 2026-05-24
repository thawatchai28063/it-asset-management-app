from __future__ import annotations

import re
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OLD_FILE = Path(r"C:\Users\it-ae\Desktop\IT ASSET TRACKING Control.xlsx")
NEW_FILE = ROOT / "IT Asset Control_repaired.xlsx"
OUT_FILE = ROOT / "reports" / "asset_tracking_comparison.xlsx"


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def excel_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        # Excel serial date, Windows 1900 date system.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
    raw = str(value).strip()
    try:
        return datetime.fromisoformat(raw.replace(" 00:00:00", "")).strftime("%Y-%m-%d")
    except ValueError:
        return raw


def normalize_status(value: str) -> str:
    raw = value.strip().lower()
    if raw == "used":
        return "in_use"
    if raw in {"not used", "not used.", "unused", ""}:
        return "available"
    if raw in {"repair", "maintenance"}:
        return "repair"
    if raw in {"retired", "scrap", "broken"}:
        return "retired"
    return raw or "available"


def normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", header.lower())


def header_map(headers: list[str]) -> dict[str, int]:
    normalized = {normalize_header(h): i for i, h in enumerate(headers)}
    aliases = {
        "asset_name": ["assetname"],
        "item_no": ["itemno", "sourceitemno"],
        "it_tag": ["ittag"],
        "employee_no": ["empno", "employeeno"],
        "assigned_user": ["name", "assigneduser"],
        "description": ["description"],
        "purchase_date": ["installed", "purchasedate"],
        "receipt_of_device": ["receiptofdevice"],
        "os_version": ["osversion"],
        "serial_number": ["serialnumber", "serialnumbersn"],
        "brand": ["band", "brand"],
        "model": ["model"],
        "asset_type": ["type", "assettype"],
        "invoice_no": ["invoiceno"],
        "date2": ["date2"],
        "vendor": ["vendor"],
        "department": ["departmentusage", "departmentusagelocation", "department"],
        "position": ["location", "position"],
        "status": ["status"],
        "point_image": ["pointimage"],
        "remark": ["remark"],
        "passwordadmin": ["passwordadmin"],
    }
    result = {}
    for key, keys in aliases.items():
        for alias in keys:
            if alias in normalized:
                result[key] = normalized[alias]
                break
    return result


def row_value(row, mapping: dict[str, int], key: str) -> str:
    index = mapping.get(key)
    if index is None or index >= len(row):
        return ""
    if key in {"purchase_date", "date2", "receipt_of_device"}:
        return excel_date(row[index])
    if key == "status":
        return normalize_status(text(row[index]))
    return text(row[index])


def read_tracking(path: Path, sheet_name: str, header_row: int) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[sheet_name]
    headers = [text(cell.value) for cell in ws[header_row]]
    mapping = header_map(headers)
    records = {}
    fields = [
        "asset_name",
        "item_no",
        "it_tag",
        "employee_no",
        "assigned_user",
        "description",
        "purchase_date",
        "receipt_of_device",
        "os_version",
        "serial_number",
        "brand",
        "model",
        "asset_type",
        "invoice_no",
        "date2",
        "vendor",
        "department",
        "position",
        "status",
        "point_image",
        "remark",
        "passwordadmin",
    ]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        tag = row_value(row, mapping, "it_tag")
        if not tag:
            continue
        records[tag] = {field: row_value(row, mapping, field) for field in fields}
    return records


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for col in range(1, ws.max_column + 1):
        width = min(42, max(12, max(len(text(ws.cell(row, col).value)) for row in range(1, min(ws.max_row, 60) + 1)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def main() -> None:
    old = read_tracking(OLD_FILE, "Assets_Import_Export", 1)
    new = read_tracking(NEW_FILE, "IT Asset Tracking", 2)

    old_tags = set(old)
    new_tags = set(new)
    added = sorted(new_tags - old_tags)
    removed = sorted(old_tags - new_tags)
    common = sorted(old_tags & new_tags)

    compare_fields = [
        "employee_no",
        "assigned_user",
        "description",
        "purchase_date",
        "os_version",
        "serial_number",
        "brand",
        "model",
        "asset_type",
        "department",
        "position",
        "status",
        "point_image",
        "remark",
        "passwordadmin",
    ]

    changed_rows = []
    for tag in common:
        diffs = []
        for field in compare_fields:
            old_value = old[tag].get(field, "")
            new_value = new[tag].get(field, "")
            if old_value != new_value:
                diffs.append((field, old_value, new_value))
        if diffs:
            changed_rows.extend((tag, field, old_value, new_value) for field, old_value, new_value in diffs)

    summary_counts = [
        ("old_total", len(old)),
        ("new_total", len(new)),
        ("same_it_tag", len(common)),
        ("added_it_tag", len(added)),
        ("removed_it_tag", len(removed)),
        ("changed_cells", len(changed_rows)),
        ("changed_assets", len(set(row[0] for row in changed_rows))),
    ]

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["metric", "value"])
    for row in summary_counts:
        ws_summary.append(row)
    ws_summary.append([])
    ws_summary.append(["new_status", "count"])
    for status, count in Counter(record["status"] for record in new.values()).most_common():
        ws_summary.append([status, count])
    style_sheet(ws_summary)

    ws_added = wb.create_sheet("Added")
    added_headers = ["it_tag"] + compare_fields
    ws_added.append(added_headers)
    for tag in added:
        ws_added.append([tag] + [new[tag].get(field, "") for field in compare_fields])
    style_sheet(ws_added)
    add_table(ws_added, "AddedAssets")

    ws_removed = wb.create_sheet("Removed")
    ws_removed.append(["it_tag"] + compare_fields)
    for tag in removed:
        ws_removed.append([tag] + [old[tag].get(field, "") for field in compare_fields])
    style_sheet(ws_removed)
    add_table(ws_removed, "RemovedAssets")

    ws_changed = wb.create_sheet("Changed")
    ws_changed.append(["it_tag", "field", "old_value", "new_value"])
    for row in changed_rows:
        ws_changed.append(row)
    style_sheet(ws_changed)
    add_table(ws_changed, "ChangedAssets")

    ws_full = wb.create_sheet("New_IT_Asset_Tracking")
    ws_full.append(["it_tag"] + compare_fields)
    for tag in sorted(new):
        ws_full.append([tag] + [new[tag].get(field, "") for field in compare_fields])
    style_sheet(ws_full)
    add_table(ws_full, "NewTracking")

    wb.save(OUT_FILE)
    print(f"old_total={len(old)}")
    print(f"new_total={len(new)}")
    print(f"added={len(added)}")
    print(f"removed={len(removed)}")
    print(f"changed_assets={len(set(row[0] for row in changed_rows))}")
    print(f"changed_cells={len(changed_rows)}")
    print(OUT_FILE)


if __name__ == "__main__":
    main()
